from psycopg2 import sql
import pandas as pd
from openpyxl import load_workbook
import psycopg2
from datetime import date
from dotenv import load_dotenv
import os
import sys

def main():
    name_file = input("Введите имя файла: \n").strip()
    data, list_exel = parser(name_file)
    print(data)
    add_db(data, list_exel)

def parser(name_file):
    try:
        wb = load_workbook(name_file, data_only=True)
        sheet = wb.active
        tabel_name = sheet.title
        
        # Проверка на пустые заголовки
        headers = [cell.value for cell in sheet[1]]
        if not any(headers):
            print("Ошибка: Первая строка (заголовки) пустая")
            sys.exit(1)
            
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Пропускаем пустые строки
                data.append(tuple(row))
                
        list_exel = {}
        for sheet in wb.worksheets[1:]:
            sheet_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Пропускаем пустые строки
                    sheet_data.append(row)
            if sheet_data:  # Добавляем только непустые листы
                list_exel[sheet.title] = sheet_data
                
        return data, list_exel
        
    except FileNotFoundError:
        print(f"Ошибка: Файл {name_file} не найден")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка при обработке файла: {str(e)}")
        sys.exit(1)

def add_db(data, list_exel):              
    try:
        load_dotenv()
        DB_URL = os.getenv("DB_URL")
        if not DB_URL:
            print("Ошибка: Не найдена переменная окружения DB_URL")
            sys.exit(1)
            
        conn = psycopg2.connect(DB_URL)
        cursor = conn.cursor()
        
        for user in data:
            try:
                if len(user) < 2:  # Проверка количества полей
                    print(f"Ошибка: Недостаточно данных в строке {user}")
                    continue
                    
                # Вставка пользователя + получение ID
                cursor.execute(
                    "INSERT INTO users (name, email) VALUES (%s, %s) RETURNING id",
                    (user[0], user[1])
                )
                user_id = cursor.fetchone()[0]

                # Обработка дополнительных параметров
                if user[0] in list_exel and list_exel[user[0]]:
                    params_data = list_exel[user[0]]
                    params_with_ids = [(user_id, *p) for p in params_data]
                    
                    cursor.executemany(
                        "INSERT INTO params (user_id, old, sex, height, weight, birthday) VALUES (%s,%s,%s,%s,%s,%s)",
                        params_with_ids
                    )

                conn.commit()
                print(f"Добавлен пользователь ID: {user_id}")

            except psycopg2.Error as e:
                conn.rollback()
                print(f"Ошибка при добавлении пользователя {user}: {e}")
    
    except psycopg2.Error as e:
        print(f"Ошибка подключения к базе данных: {e}")
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

if __name__ == "__main__":
    main()