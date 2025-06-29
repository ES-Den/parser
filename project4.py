from psycopg2 import sql
import pandas as pd
from openpyxl import load_workbook
import psycopg2
from datetime import date
from dotenv import load_dotenv
import os


name_file = input("Введите имя файла: \n")

def parser(name_file):
    try:
        wb = load_workbook(name_file, data_only=True)
        sheet = wb.active
        tabel_name = wb.active.title
    except Exception as e:
        print(print(f"Ошибка: Не удалось загрузить файл {name_file}. Причина: {e}"))
        exit()
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append(tuple(row))
    list_exel = dict()


    for sheet in wb.worksheets[1:]:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            list_exel[sheet.title] = list_exel.get(sheet.title, []) + [row]
    return data, list_exel
    
data, list_exel = parser(name_file)
print(data)


def add_db(parser):
    data, list_exel = parser              
    load_dotenv()
    DB_URL = os.getenv("DB_URL")
    conn = psycopg2.connect(DB_URL)
    cursor = conn.cursor()
    for user in data:
        user_data = user
        params_data =list_exel[user[0]]
        try:
            # Вставка пользователя + получение ID
            cursor.execute(
                "INSERT INTO users (name, email) VALUES (%s, %s) RETURNING id",
                user_data
            )
            user_id = cursor.fetchone()[0]  # Получаем сгенерированный ID

            if params_data:
                params_with_ids = [(user_id, *p) for p in params_data]
                cursor.executemany(
                "INSERT INTO params (user_id, old, sex, height, weight, birthday) VALUES (%s,%s,%s,%s,%s,%s)",
                params_with_ids
                )

            conn.commit()
            print(f"Добавлен пользователь ID: {user_id}")

        except Exception as e:
            conn.rollback()
            print(f"Ошибка: {e}")   
    
    cursor.close()
    conn.close()


add_db(parser(name_file))



     